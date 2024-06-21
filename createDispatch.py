from openpyxl import Workbook, load_workbook
from operator import itemgetter
from itertools import groupby
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment


def createDispatch(formattedSolution, filename):
    print("opening workbook in createDispatch")
    wb = load_workbook(filename)
    dispatchSheet = wb.create_sheet("Dispatch")
    dispatchSheet.column_dimensions["A"].width = 11
    dispatchSheet.column_dimensions["B"].width = 9
    dispatchSheet.column_dimensions["C"].width = 14
    dispatchSheet.column_dimensions["D"].width = 11
    dispatchSheet.column_dimensions["E"].width = 15
    dispatchSheet.column_dimensions["F"].width = 13
    dispatchSheet.column_dimensions["G"].width = 11
    dispatchSheet.column_dimensions["H"].width = 18
    dispatchSheet.column_dimensions["I"].width = 40

    driverView = wb.create_sheet("Driver View")
    driverView.column_dimensions["A"].width = 11
    driverView.column_dimensions["B"].width = 9
    driverView.column_dimensions["C"].width = 14
    driverView.column_dimensions["D"].width = 11
    driverView.column_dimensions["E"].width = 15
    driverView.column_dimensions["F"].width = 13
    driverView.column_dimensions["G"].width = 11
    driverView.column_dimensions["H"].width = 18
    driverView.column_dimensions["I"].width = 40

    dispatchSheet.merge_cells("A1:J1")
    dispatchSheet.merge_cells("A6:J6")
    dispatchSheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
    dispatchSheet["A1"].value = "Today's Date goes here"
    dispatchSheet["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    dispatchSheet.merge_cells("A2:F2")
    dispatchSheet.merge_cells("A3:F3")
    dispatchSheet.merge_cells("A4:F4")
    dispatchSheet.merge_cells("A5:F5")
    dispatchSheet["A2"].alignment = Alignment(horizontal='center', vertical='center')
    dispatchSheet["A2"].value = "Ship - Port Call - AA - Dock"

    dispatchSheet["G2"].alignment = Alignment(horizontal='center', vertical='center')
    dispatchSheet["G2"].value = "Dock Rep"
    dispatchSheet["H2"].alignment = Alignment(horizontal='center', vertical='center')
    dispatchSheet["H2"].value = "Contact"
    dispatchSheet["I2"].alignment = Alignment(horizontal='center', vertical='center')
    dispatchSheet["I2"].value = "Mod & Standby"

    trips = []
    dispatchSheet.append(["Driver", "Vehicle #", "Pick up Location", "Pick Up Time", "Drop Off Location", "Drop Off Time", "Passengers", "Ship", "Tour Name", "Dock Rep"])
    driverView.append(["Driver", "Vehicle #", "Pick up Location", "Pick Up Time", "Drop Off Location", "Drop Off Time", "Passengers", "Ship", "Tour Name", "Dock Rep"])

    for route in formattedSolution:
        i = 0
        while i < len(route["pickupLocations"]):
            trips.append([route["driver"], route["vehicle"], route["pickupLocations"][i], route["pickupTimes"][i], route["dropoffLocations"][i], route["dropoffTimes"][i], route["passengers"][i], route["shipName"][i], route["tourName"][i], route["dockRep"][i]])
            i += 1

    sorted_trips = sorted(trips, key=lambda x: x[3])
    for trip in sorted_trips:
        dispatchSheet.append(trip)

    groupedByVehicle = groupby(sorted(trips, key=itemgetter(1)), key=itemgetter(1))
    
    sortedGroupedByVehicle = []
    for key, group in groupedByVehicle:
        sortedGroup = sorted(group, key=itemgetter(1))
        sortedGroupedByVehicle.extend(sortedGroup)
    
    for trip in sortedGroupedByVehicle:
        driverView.append(trip)

    for cell in dispatchSheet["D"]:
        cell.number_format = "hh:mm"
    for cell in dispatchSheet["F"]:
        cell.number_format = "hh:mm"
    for cell in driverView["D"]:
        cell.number_format = "hh:mm"
    for cell in driverView["F"]:
        cell.number_format = "hh:mm"

    for cell in dispatchSheet[2]:
        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    for cell in driverView[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
        cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
        
    for cell in dispatchSheet[7]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
        cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    wb.save(filename)
    wb.close()
    
