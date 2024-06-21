from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


vehicleNumbers = [1, 2, 4, 6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20, 21, 22, 23, 24, 25, 27, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66]

vehicleCapacities = [20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 11, 14, 14, 14, 3, 14, 14, 14, 14, 14, 14, 14, 14, 14, 11, 11, 14, 28, 2, 5, 14, 35, 35, 35, 35, 32, 24, 14, 14, 14, 14, 14, 14, 20, 20, 20, 20, 20, 20, 20, 20, 3, 50, 50, 50, 50, 50, 50, 50, 50]


def generateTemplate(excelFile):
    wb = load_workbook(excelFile)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print("sheet must be empty")
                    return False
    
    for sheet in wb.worksheets:
        wb.remove(sheet)
    
    DataSheet = wb.create_sheet("Data Sheet")
    ToursAndLocations = wb.create_sheet("Tours and Locations")
    DistanceMatrix = wb.create_sheet("Distance Matrix")

    #Data Sheet 
    DataSheet["A1"].value = "Ship Name"
    DataSheet["B1"].value = "Tour Name"
    DataSheet["C1"].value = "return? y/n"
    DataSheet["D1"].value = "pick up location"
    DataSheet["E1"].value = "drop off location"
    DataSheet["F1"].value = "pick up time"
    DataSheet["G1"].value = "drop off time"
    DataSheet["H1"].value = "passenger count"
    DataSheet["J1"].value = "ship names"
    DataSheet["K1"].value = "ship locations"
    DataSheet["L1"].value = "dock reps"
    DataSheet["N1"].value = "vehicle numbers"
    DataSheet["O1"].value = "vehicle capacities"
    DataSheet["P1"].value = "out of service? y/n"
    DataSheet["Q1"].value = "drivers"
    DataSheet["R1"].value = "base"
    DataSheet["R2"].value = "Cambria"

    DataSheet.column_dimensions["A"].width = 20
    DataSheet.column_dimensions["B"].width = 38
    DataSheet.column_dimensions["C"].width = 12
    DataSheet.column_dimensions["D"].width = 20
    DataSheet.column_dimensions["E"].width = 21
    DataSheet.column_dimensions["F"].width = 12
    DataSheet.column_dimensions["G"].width = 13
    DataSheet.column_dimensions["H"].width = 14
    DataSheet.column_dimensions["I"].width = 10
    DataSheet.column_dimensions["J"].width = 17
    DataSheet.column_dimensions["K"].width = 14
    DataSheet.column_dimensions["L"].width = 13
    DataSheet.column_dimensions["M"].width = 10
    DataSheet.column_dimensions["N"].width = 15
    DataSheet.column_dimensions["O"].width = 15
    DataSheet.column_dimensions["P"].width = 16
    DataSheet.column_dimensions["Q"].width = 15
    DataSheet.column_dimensions["R"].width = 8

    DataSheet["A1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["B1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["C1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["F1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["G1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["H1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["J1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["K1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["L1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["P1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["Q1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["D1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["E1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["N1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["O1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["R1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")

    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    for col in DataSheet.iter_cols():
        for cell in col:
            cell.border = border
            break



    counter = 2
    
    while counter < 200:
        DataSheet[f"D{counter}"].value = f"=IF(C{counter}=\"y\", IFERROR(INDEX(\'Tours and Locations\'!$B$1:$B$100, MATCH(B{counter}, \'Tours and Locations\'!$A$1:$A$100, 0)), \"\"), IFERROR(INDEX($K$1:$K$100, MATCH(A{counter}, $J$1:$J$100, 0)), \"\") )"
        DataSheet[f"E{counter}"].value = f"=IF(C{counter}<>\"y\", IFERROR(INDEX(\'Tours and Locations\'!$B$1:$B$100, MATCH(B{counter}, \'Tours and Locations\'!$A$1:$A$100, 0)), \"\"), IFERROR(INDEX($K$1:$K$100, MATCH(A{counter}, $J$1:$J$100, 0)), \"\"))"

        DataSheet[f"F{counter}"].value = f"=IF(C{counter}=\"y\", IFERROR(G{counter} - TIME(0, IFERROR(INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(D{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(E{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0)), INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(E{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(D{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0))), 0), \"\"), \"\")"
        DataSheet[f"F{counter}"].number_format = 'hh:mm'
        DataSheet[f"G{counter}"].value = f"=IF(C{counter}<>\"y\", IFERROR(F{counter} + TIME(0, IFERROR(INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(D{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(E{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0)), INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(E{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(D{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0))), 0), \"\"), \"\")"
        DataSheet[f"G{counter}"].number_format = 'hh:mm'
        counter += 1

    counter = 2
    while counter <= len(vehicleNumbers):
        DataSheet[f"N{counter}"].value = vehicleNumbers[counter-2]
        DataSheet[f"O{counter}"].value = vehicleCapacities[counter-2]
        DataSheet[f"P{counter}"].value = "y"
        counter += 1

    #Tours and Locations
    ToursAndLocationsData = [["Tour Names", "Drop off Locations"], ["Adventure Karts - Whipple Creek", "Whipple Creek"], ["Aerial Zip & Rappel - Wood Rd.", "Wood Road"],
    ["AK Duck Shuttle", "Front Street Extension"], ["AK Family Fun Exclusive", "Front Street Extension"], 
    ["AK First City Highlights & LJ Show", "Front Street Extension"], ["AK Lodge Adv & Seafeast - Clover Pass Resort", "Clover Pass"], ["Aleutian Ballad - Dock 3", "Dock 3"], 
    ["Baranof Fishing Excursions - Disco Center", "Disco"], ["Best of Ketchikan George Inlet Cannery", "Cannery"], 
    ["Cannery, Crab & Beer Flight - Cannery", "Cannery"], ["Cannery, Crab & Beer Flight - GIL", "George Inlet"], 
    ["Cannery, Small Bites & Saxman - Cannery", "Cannery"], ["Cannery, Small Bites & Saxman - Saxman", "Saxman"], 
    ["Canoes & Nature Trail - Harriet Hunt Lake", "Harriot Hunt"], ["Crab Cake Expedition - GIL", "George Inlet"], 
    ["Culinary, Cultural & LJ Show", "Saxman"], ["Eagle Island Kayaks - Potter Rd.", "Potter Road"], 
    ["George Inlet Lodge - Crab Feast", "George Inlet"], ["GIL Brunch & Saxman Showcase - GIL", "George Inlet"], 
    ["GIL Brunch & Saxman Showcase - Saxman", "Saxman"], ["Historic Cannery & LJ Show - Cannery", "Cannery"], 
    ["Historic Cannery & LJ Show - Disco Center", "Disco"], ["Historic Cannery & LJ Show - Front St. or Dock 1", "Dock 1"], 
    ["Hovercraft Shuttle", "Dock 3"], ["Jeeps & Canoes - Harriet Hunt Lake", "Harriot Hunt"], 
    ["Jeeps & Canoes - Warehouse", "Jeep Base"], ["Ketchikan Highlights by Trolley", "Saxman"], ["KIA - Gravina Side", "Airport"], 
    ["KIA - Ketchikan Side", "Airport"], ["Knudson Cove Sportfishing - Knudson Cove Marina", "Knudson"], 
    ["KTN Native Walking Tour - CFL", "Cape Fox Lodge"], ["KTN Native Walking Tour - Dock 4", "Dock 4"], 
    ["KTN Native Walking Tour - Salmon Ladder", "Salmon Ladder"], ["KTN Trolley & LJ Show", "Front Street Extension"], 
    ["Lighthouse, Eagles & Totems - Ward Cove", "Lighthouse"], ["LJ Show & Crab Feast - Front St. or Dock 1", "Dock 1"], 
    ["LJ Show &Crab Feast - GIL", "George Inlet"], ["Lumberjack Show - Disco Center", "Disco"], 
    ["Misty Fjords Flightseeing - Taquan Air", "Taquan"], ["Neets Bay Bear Cruise - Clover Pass Resort", "Clover Pass"], 
    ["Out to Sea Expedition - Thomas St.", "Thomas Street"], ["Rainforest Island Adv. - Knudson Cove Marina", "Knudson"], 
    ["RF Sanctuary Walk - Wood Rd.", "Wood Road"], ["RF Walk & Crab Feast - ARS to GIL", "George Inlet"], 
    ["RF Walk & Crab Feast - GIL", "George Inlet"], ["RF Walk & Crab Feast - GIL to ARS", "George Inlet"], 
    ["RF Walk & Crab Feast - Wood Rd.", "Wood Road"], ["RF Zip & Skybridge - Wood Rd.", "Wood Road"], 
    ["Ropes & Challenge Zips - Potter Rd.", "Potter Road"], ["Salmon Dip Making & Photostop", "Saxman"], 
    ["Salmon Dip Making & Photostop - Parking Lot", "Saxman"], ["Sawmill & Crab Feast - GIL", "George Inlet"], 
    ["Sawmill & Crab Feast - Parking Lot", "Saxman"], ["Sawmill & Crab Feast - Sawmill", "Saxman"], 
    ["Sawmill, Totems, & Discovery Center - Disco Center", "Disco"], ["Sawmill, Totems, & Discovery Center - Parking Lot", "Saxman"], 
    ["Sawmill, Totems, & Discovery Center - Sawmill", "Saxman"], ["Saxman Craft Showcase - Community Center", "Saxman"], 
    ["Saxman Craft Showcase - Parking Lot", "Saxman"], ["Saxman Native Village", "Saxman"], ["Saxman Native Village & LJ Show", "Saxman"], 
    ["Saxman Photostop & Crab Feast - GIL", "George Inlet"], ["Saxman Photostop & Crab Feast - Saxman", "Saxman"], 
    ["Tatoosh Island Kayaks - Potter Rd.", "Potter Road"], ["UTV Safari & GIL Crab Meal - GIL", "George Inlet"], 
    ["UTV Safari & GIL Crab Meal - White River", "White River"], ["White River UTV Base Camp", "White River"], 
    ["Wilderness Zodiac Quest - Cannery", "Cannery"], ["Wildlife Bear Country - Wood Rd.", "Wood Road"]]
    
    for nameAndLocation in ToursAndLocationsData:
        ToursAndLocations.append(nameAndLocation)

    ToursAndLocations.column_dimensions["A"].width = 43
    ToursAndLocations.column_dimensions["B"].width = 26
    ToursAndLocations["A1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    ToursAndLocations["B1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")

    #Distance Matrix
    DistanceMatrixData = [["Distance Matrix", "Ward Cove", "Dock 1", "Dock 2", "Dock  3", "Dock  4", "Front Street Extension", "George Inlet", "Cannery", "Wood Road", "Cape Fox Lodge", "Saxman", "Disco", "Taquan", "Jeep Base", "Harriot Hunt", "White River", "Whipple Creek", "Knudson", "Clover Pass", "Totem Bight", "Lighthouse", "Cambria", "Potter Road", "Airport", "Salmon Ladder", "Thomas Street"],
    ["Ward Cove",	0,	25,	25,	25,	25,	25,	50,	60,	40,	30,	35,	25,	20,	10,	30,	45,	25,	30,	40,	25,	5,	20,	40,	20,	30,	25],
    ["Dock 1", 25, 0,	5,	5,	10,	5,	40,	50,	25,	15,	25,	5,	10,	25,	50,	55,	40,	45,	45,	35,	25,	15,	45,	10,	15,	5],
    ["Dock 2", 25, 5,	0,	5,	10,	5,	40,	50,	25,	15,	25,	5,	10,	25,	50,	55,	40,	45,	45,	35,	25,	15,	45,	10,	15,	5],
    ["Dock 3", 25, 5,	5, 0, 10, 5,	40,	50,	25,	15,	25,	5,	10,	25,	50,	55,	40,	45,	45,	35,	20,	15,	45,	10,	15,	5],
    ["Dock 4", 25, 10, 10, 10, 0,	8,	45,	50,	30,	20,	30,	8,	10,	20,	45,	55,	35,	40,	45,	35,	20,	15,	45,	10,	20,	8],
    ["Front Street Extension",	25,	5,	5,	5,	8,	0,	40,	50,	25,	15,	25,	5,	10,	25,	50,	55,	40,	45,	45,	35,	25,	15,	45,	10,	15,	5,],
    ["George Inlet",	50,	40,	40,	40,	45,	40,	0,	10,	20,	45,	25,	40,	45,	50,	70,	85,	70,	70,	70,	60,	50,	50,	70,	45,	45,	40],
    ["Cannery",	60,	50,	50,	50,	50,	50,	10,	0,	25,	45,	30,	50,	55,	60,	80,	95,	80,	80,	80,	70,	60,	55,	80,	55,	45,	50],
    ["Wood Road",	40,	25,	25,	25,	30,	25,	20,	25,	0,	30,	20,	30,	35,	40,	65,	70,	60,	70,	75,	60,	45,	35,	75,	35,	30,	30],
    ["Cape Fox Lodge",	30,	15,	15,	15,	20,	15,	45,	45,	30,	0,	30,	15,	20,	30,	55,	60,	45,	50,	55,	40,	30,	25,	55,	20,	5,	15],
    ["Saxman",	35,	25,	25,	25,	30,	25,	25,	30,	20,	30,	0,	25,	30,	40,	60,	70,	50,	50,	55,	45,	55,	30,	55,	30,	30,	25],
    ["Disco",	25,	5,	5,	5,	8,	5,	40,	50,	30,	15,	25,	0,	10,	25,	50,	55,	40,	45,	45,	35,	25,	20,	45,	10,	15,	5],
    ["Taquan",	20,	10,	10,	10,	10,	10,	45,	55,	35,	20,	30,	10,	0,	15,	45,	50,	35,	35,	40,	30,	20,	5,	40,	5,	20,	10],
    ["Jeep Base",	10,	25,	25,	25,	20,	25,	50,	60,	40,	30,	40,	25,	15,	0,	30,	45,	30,	30,	35,	25,	10,	15,	35,	15,	30,	25],
    ["Harriot Hunt",	30,	50,	50,	50,	45,	50,	70,	80,	65,	55,	60,	50,	45,	30,	0,	20,	40,	50,	50,	40,	30,	45,	50,	45,	55,	50],
    ["White River",	45,	55,	55,	55,	55,	55,	85,	95,	70,	60,	70,	55,	50,	45,	20,	0,	50,	60,	60,	50,	45,	50,	60,	50,	60,	55],
    ["Whipple Creek",	25,	40,	40,	40,	35,	40,	70,	80,	60,	45,	50,	40,	35,	30,	40,	50,	0,	30,	40,	25,	25,	35,	40,	35,	45,	40],
    ["Knudson",	30,	45,	45,	45,	40,	45,	70,	80,	70,	50,	50,	45,	35,	30,	50,	60,	30,	0,	10,	20,	30,	35,	10,	35,	50,	45],
    ["Clover Pass",	40,	45,	45,	45,	45,	45,	70,	80,	75,	55,	55,	45,	40,	35,	50,	60,	40,	10,	0,	25,	30,	40,	10,	40,	55,	45],
    ["Totem Bight",	25,	35,	35,	35,	35,	35,	60,	70,	60,	40,	45,	35,	30,	25,	40,	50,	25,	20,	25,	0,	20,	30,	25,	30,	40,	35],
    ["Lighthouse",	5,	25,	25,	20,	20,	25,	50,	60,	45,	30,	55,	25,	20,	10,	30,	45,	25,	30,	30,	20,	0,	25,	30,	20,	30,	25],
    ["Cambria",	20,	15,	15,	15,	15,	15,	50,	55,	35,	25,	30,	20,	5, 15,	45,	50,	35,	35,	40,	30,	25,	0,	40,	5,	25,	20],
    ["Potter Road",	40,	45,	45,	45,	45,	45,	70,	80,	75,	55,	55,	45,	40,	35,	50,	60,	40,	10,	10,	25,	30,	40,	0,	40,	55,	45],
    ["Airport",	20,	10,	10,	10,	10,	10,	45,	55,	35,	20,	30,	10,	5,	15,	45,	50,	35,	35,	40,	30,	20,	5,	40,	0,	20,	10],
    ["Salmon Ladder",	30,	15,	15,	15,	20,	15,	45,	45,	30,	5,	30,	15,	20,	30,	55,	60,	45,	50,	55,	40,	30,	25,	55,	20,	0,	15],
    ["Thomas Street",	25,	5,	5,	5,	8,	5,	40,	50,	30,	15,	25,	5,	10,	25,	50,	55,	40,	45,	45,	35,	25,	20,	45,	15,	15,	0]
    ]

    for line in DistanceMatrixData:
        DistanceMatrix.append(line)
    
    


    
    
    wb.save(excelFile)

    print("Thank you for providing an empty sheet.")
