import sqlite3
from openpyxl import load_workbook
from loadTemplate import loadTemplate

def create_connection(db_file):

    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print("connected")
    except sqlite3.Error as e:
        print(e)
    return conn
    

def create_table(conn):

    print("about to create table")
    create_table_sql = """CREATE TABLE IF NOT EXISTS templateModel 
    (id INTEGER PRIMARY KEY AUTOINCREMENT,
    dataSheet TEXT,
    toursAndLocations TEXT,
    distanceMatrix TEXT,
    name TEXT
    )"""

    try:
        cursor = conn.cursor()
        cursor.execute(create_table_sql)
        print("table created")
    except sqlite3.Error as e:
        print(e)
    
def insert(conn, data):
    InsertSql = """INSERT INTO templateModel(dataSheet, toursAndLocations, distanceMatrix, name)
    VALUES(?, ?, ?, ?)"""

    cursor = conn.cursor()
    cursor.execute(InsertSql, data)
    conn.commit()
    print("data added")
    return cursor.lastrowid

def deleteTemplate(conn, id):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM templateModel WHERE id=?", (id,))



#with conn:
    #create_table(conn)

    #insert(conn, ['1,2,3', '4,5,6', '7,8,9'])


def getWorksheetFromExcel(wb, workSheetName):
    dataList = []
    countSinceData = 0
    for column in wb[workSheetName].iter_cols():
        dataList.append([])
        for cell in column:
            if cell.value != None:
                countSinceData = 0
            elif cell.value == None:
                countSinceData += 1
            
            dataList[-1].append(cell.value)
        if countSinceData > 3:
            dataList[-1] = dataList[-1][:1-countSinceData]        
        dataList[-1].append(None)

    dataString = ""
    for column in dataList:
        dataString += "@"
        for cell in column:
            dataString += f"{cell}"
            dataString += "~"
        dataString = dataString[1:-1]
    return dataString


def saveTemplate(excelFile, templateName):

    wb = load_workbook(excelFile)
    with conn:
        dataSheetString = getWorksheetFromExcel(wb, "Data Sheet")
        ToursAndLocationsString = getWorksheetFromExcel(wb, "Tours and Locations")
        DistanceMatrixString = getWorksheetFromExcel(wb, "Distance Matrix")
        insert(conn, [dataSheetString, ToursAndLocationsString, DistanceMatrixString, templateName])



#conn = create_connection("test.db")
#with conn:
    #create_table(conn)
    #saveTemplate("/Users/user-1/Desktop/TestDispatchGenerator.xlsx", "firstTemplate")
