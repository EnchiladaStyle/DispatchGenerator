import sqlite3
from openpyxl import load_workbook
from styleTemplate import styleTemplate

def selectTemplates(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM templateModel")
    conn.commit()

    rows = cursor.fetchall()
    return rows[0]

def create_connection(db_file):

    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print("connected")
    except sqlite3.Error as e:
        print(e)
    return conn

def getTemplateFromSQL(conn, id):
    cursor = conn.cursor()
    cursor.execute("SELECT dataSheet, toursAndLocations, distanceMatrix FROM templateModel WHERE id = ?", [id,])
    conn.commit()

    rows = cursor.fetchall()
    return rows[0]

def parseStringToLists(dataString):
    
    solution = []

    for i in range(len(dataString.split("@"))):
        solution.append(dataString.split("@")[i].split("~"))

    return solution

def populateWorkSheet(sheet, newDataList):

    for col_idx, col_data in enumerate(newDataList, start=1):
        for row_idx, value in enumerate(col_data, start=1):
            if value == "None":
                sheet.cell(row=row_idx, column=col_idx, value=None)
            else:
                sheet.cell(row=row_idx, column=col_idx, value=value)


def loadTemplate(excelFile):

    wb = load_workbook(excelFile)
    conn = create_connection("test.db")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print("sheet must be empty")
                    return False
    
    for sheet in wb.worksheets:
        wb.remove(sheet)

    with conn:
        workSheetStrings = getTemplateFromSQL(conn, 1)

    newDataSheetList = parseStringToLists(workSheetStrings[0])
    newToursAndLocationsList = parseStringToLists(workSheetStrings[1])
    newDistanceMatrixList = parseStringToLists(workSheetStrings[2])
    
    DataSheet = wb.create_sheet("Data Sheet")
    ToursAndLocations = wb.create_sheet("Tours and Locations")
    DistanceMatrix = wb.create_sheet("Distance Matrix")

    populateWorkSheet(DataSheet, newDataSheetList)
    populateWorkSheet(ToursAndLocations, newToursAndLocationsList)
    populateWorkSheet(DistanceMatrix, newDistanceMatrixList)

    wb.save(excelFile)
    wb.close()

    styleTemplate(excelFile)

loadTemplate("/Users/user-1/Desktop/BlankWorkbook.xlsx")
    